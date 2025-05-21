import openpyxl

def Read_Data_From_Excel(file, sheet_name, row_no, col_no):
    excel_file = openpyxl.load_workbook(file)
    sheet = excel_file[sheet_name]
    return sheet.cell(row=row_no, column=col_no).value

def Write_Data_To_Excel(file, sheet_name, row_no, col_no, data):
    excel_file = openpyxl.load_workbook(file)
    sheet = excel_file[sheet_name]
    sheet.cell(row=row_no, column=col_no).value = data
    excel_file.save(file)

def Max_Row_Count_From_Excel(file, sheet_name):
    excel_file = openpyxl.load_workbook(file)
    sheet = excel_file[sheet_name]
    return sheet.max_row
